import openpyxl
import os
from openpyxl import load_workbook

def hyperlink_excell_col_values(excell_file_path,row_nr, col_nr, hyperlink_base):
    # Wczytaj plik Excel i wybierz aktywny arkusz
    workbook = load_workbook(excell_file_path)
    sheet = workbook.active

    # Zakładając, że ścieżki są w col_nr kolumnie, iteruj przez wiersze od row_nr
    for row in sheet.iter_rows(min_row=row_nr, min_col=col_nr, max_col=col_nr, max_row=sheet.max_row):
        for cell in row:
            # Sprawdź, czy komórka zawiera ścieżkę
            if cell.value and isinstance(cell.value, str):
                # Utwórz hiperłącze, dodając prefiks
                cell.hyperlink = hyperlink_base + cell.value.replace("\\", "/")
                # Opcjonalnie, możesz ustawić wyświetlany tekst hiperłącza na samą ścieżkę
                cell.value = cell.value

    # Zapisz plik Excel z hiperłączami
    edited_excell = os.path.join(os.path.dirname(excell_file_path),os.path.basename(excell_file_path).replace('.xlsx','_hyperlinked.xlsx'))
    workbook.save(edited_excell)

    print("Hiperłącza zostały dodane do ścieżek w pliku Excel i zapisane jako '{}'.".format(edited_excell))

if __name__ == '__main__':
    excell_path = r"C:\robo\Przeglad_LOKAL_ROBO\KOMPLEKSOWA_OCENA\3M.1.06a Zestawienie uwag 20240719 v2.01.xlsx"
    numer_wiersza = 4
    numer_kolumny = 7
    trzon_hiperlacza = r'https://arcadiso365.sharepoint.com/teams/ch-10537335/Shared Documents/External/PRODUKTY DO KONTROLI/3M PRZEGLAD MZPIMRP/08 PRODUKTY/20240719/TABELE/'

    hyperlink_excell_col_values(excell_path, numer_wiersza, numer_kolumny,trzon_hiperlacza)
#!/usr/bin/env python
#-*- coding: utf-8 -*-

import os
import arcpy
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
from openpyxl.chart import ScatterChart, Reference, Series

def styleHeadline():
    hstyle = NamedStyle(name="headline")
    hstyle.font = Font(bold=True, size=11)
    bd = Side(style='thin', color="000000")
    hstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    hstyle.fill = PatternFill("solid", fgColor="00C0C0C0")
    hstyle.alignment = Alignment(horizontal="center", vertical="center")
    return hstyle

def styleStandart():
    sstyle = NamedStyle(name="standart")
    sstyle.font = Font(bold=False, size=11)
    bd = Side(style='thin', color="000000")
    sstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sstyle.alignment = Alignment(horizontal="center", vertical="center")
    return sstyle

def createTemplateXlsx(folder):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('PUNKTY_PRZEKROJOW_DOLINOWYCH', 0)
    ws['A1'] = 'Nazwa cieku'
    ws['B1'] = 'Identyfikator hydrograficzny'
    ws['C1'] = 'Numer przekroju i punktu pomiarowego'
    ws['D1'] = 'Współrzędna X [m]'
    ws['E1'] = 'Współrzędna Y [m]'
    ws['F1'] = 'Z - rzędna'+'\n'+'[m n.p.m.]'
    ws['G1'] = 'Odległość [m]'
    ws['H1'] = 'Kod punktu'
    ws['I1'] = 'Kod formy pokrycia terenu'
    ws['J1'] = 'Rzędna zw. wody'+'\n'+'[m n.p.m.]'
    ws['K1'] = 'Data pomiaru'
    ws['L1'] = 'Numer fotografii'
    ws['M1'] = 'Uwagi'

    headlinestyle = styleHeadline()
    for m in range(1, 14):
        try:
            ws.cell(row=1, column=m).style=headlinestyle
        except:
            ws.cell(row=1, column=m).style="headline"

    output_xlsx = 'template.xlsx'
    wb.save(os.path.join(folder, output_xlsx))
    wb.close()
    return output_xlsx

def pktPrzekroi(lPrzekroje, lPomiary, lTerasy, tolerancja):

    # wyselekcjonuj
    lsPrzekroje = "in_memory/lsPrzekroje"
    arcpy.Select_analysis(lPrzekroje, lsPrzekroje)
    lsPomiary = "in_memory/lsPomiary"
    arcpy.Select_analysis(lPomiary, lsPomiary)
    lsTerasy = "in_memory/lsTerasy"
    arcpy.Select_analysis(lTerasy, lsTerasy)
    outTblpom = "in_memory/tabPomiar"
    outTblter = "in_memory/tabTerasy"
    mergedTab = "in_memory/tabMerge"
    out_dataset = "in_memory/tabOutput"

    # przygotowanie route
    arcpy.AddMessage("1 - Przygotowanie routes...")
    arcpy.AddField_management(lsPrzekroje, "rtStart", "FLOAT", 10, 3)
    arcpy.AddField_management(lsPrzekroje, "rtEnd", "FLOAT", 10, 3)
    fldKm = "NR_PRZEK"

    arcpy.CalculateField_management(lsPrzekroje, "rtStart", "0.0", "PYTHON")
    arcpy.CalculateField_management(lsPrzekroje, "rtEnd", "!shape.length!", "PYTHON")
    rtPrzekroje = "in_memory/rtPrzekroje"
    arcpy.CreateRoutes_lr(lsPrzekroje, fldKm, rtPrzekroje, "TWO_FIELDS", "rtStart", "rtEnd")

    arcpy.AddMessage("2 - Kilometracja punktow przekroi...")

    arcpy.LocateFeaturesAlongRoutes_lr(lsPomiary, rtPrzekroje, fldKm, tolerancja, outTblpom)
    arcpy.LocateFeaturesAlongRoutes_lr(lsTerasy, rtPrzekroje, fldKm, tolerancja, outTblter)
    expr1 = "round(!MEAS!,2)"
    arcpy.CalculateField_management(outTblpom, 'MEAS', expr1, "PYTHON")
    expr2= "round(!MEAS!)"
    arcpy.CalculateField_management(outTblter, 'MEAS', expr2, "PYTHON")

    arcpy.AddMessage("3 - Przygotowanie tabel z punktami przekroi...")

    przKorDict = {}
    with arcpy.da.SearchCursor(outTblpom, ['NR_PRZEK', 'MEAS']) as cur:
        for row in cur:
            if row[0] in przKorDict:
                if row[1] < przKorDict[row[0]][0]:
                    przKorDict[row[0]][0] = row[1]
                if row[1] > przKorDict[row[0]][1]:
                    przKorDict[row[0]][1] = row[1]
            else:
                przKorDict[row[0]] = [row[1], row[1]]
    del cur

    with arcpy.da.UpdateCursor(outTblter, ['NR_PRZEK', 'MEAS']) as cur:
        for row in cur:
            if row[1] > przKorDict[row[0]][0] and row[1] < przKorDict[row[0]][1]:
                cur.deleteRow()
    del cur

    arcpy.AddMessage("4 - Opracowanie tabel wynikowych...")

    arcpy.Merge_management([outTblter, outTblpom], mergedTab)
    sort_fields = [["NR_PRZEK", "ASCENDING"], ["MEAS", "ASCENDING"]]
    arcpy.Sort_management(mergedTab, out_dataset, sort_fields)

    return out_dataset

def addScatterChart(worsheet, title, label, xCol, yCol, minRow, maxRow):
    chart = ScatterChart()
    chart.title = title
    chart.style = 1
    chart.y_axis.title = u'Z - rzędna [m n.p.m.]'
    chart.x_axis.title = u'Długość [m]'

    xData = Reference(worsheet, min_col=xCol, min_row=minRow, max_row=maxRow)
    yData = Reference(worsheet, min_col=yCol, min_row=minRow, max_row=maxRow)
    seria = Series(yData, xData, title = label)
    chart.series.append(seria)
    cell = "O"+str(minRow)
    worsheet.add_chart(chart, cell)
    return

def insertXlsxData(shpDataTab, folder, outputXlsx):
    print("5 - Zapisywanie danych przekroi do xlsx...")
    filename = createTemplateXlsx(folder)
    stylest=styleStandart()
    wb = openpyxl.load_workbook(os.path.join(folder, filename))
    ws = wb['PUNKTY_PRZEKROJOW_DOLINOWYCH']
    n = 1
    r = 1
    with arcpy.da.SearchCursor(shpDataTab, ['MEAS', 'NAZWA', 'ID_HYD_R', 'X', 'Y', 'Z', 'KOD_PKT', 'KOD_TERENU', 'RZEDNA_ZW', 'DATA_POM', 'FOTO', 'NR_PRZEK', 'NAZWA_RZGW', 'UWAGI']) as cur:
        for row in cur:
            if row[0] == 1:
                nr_pkt_l = 10001
                nr_pkt = 1
                nr_pkt_p = 20001
                minRowOst = n + 1
                # nazwa cieku
                ws.cell(row=n + 1, column=1, value=row[12])
                # id_hyd
                ws.cell(row=n + 1, column=2, value=row[2])

            if row[0] == 1 and n != 1:
                addScatterChart(ws, chartTitle, label, 7, 6, n+2-r, n)
                r = 1
            r += 1
            label = row[11]
            chartTitle = u"Przekrój dolinowy nr " + row[11]
            # Numer przekroju i punktu pomiarowego
            if row[1] is None and nr_pkt == 1:
                nr = row[11]+'.'+str(nr_pkt_l)
                ws.cell(row=n + 1, column=3, value = nr)
                nr_pkt_l += 1

            elif row[1] is not None and row[0] != 1:
                if nr_pkt < 10:
                    nr = row[11] + '.0' + str(nr_pkt)
                else:
                    nr = row[11] + '.' + str(nr_pkt)
                ws.cell(row=n + 1, column=3, value = nr)
                nr_pkt += 1

            elif row[1] is None and row[0] != 1 and nr_pkt != 1:
                nr = row[11]+'.'+str(nr_pkt_p)
                ws.cell(row=n + 1, column=3, value = nr)
                nr_pkt_p += 1

            #X  #Y
            if row[1] is None:
                ws.cell(row=n + 1, column=4, value=round(row[4], 2))
                ws.cell(row=n + 1, column=5, value=round(row[3], 2))
            else:
                ws.cell(row=n + 1, column=4, value=round(row[3], 2))
                ws.cell(row=n + 1, column=5, value=round(row[4], 2))

            #Z
            ws.cell(row=n + 1, column=6, value=round(row[5], 2))
            #Meas
            ws.cell(row=n + 1, column=7, value=row[0])
            #kod punktu
            ws.cell(row=n + 1, column=8, value=row[6])
            #kod formy pokrycia terenu
            ws.cell(row=n + 1, column=9, value=row[7])
            #Rzedna zw
            if row[8] != 0:
                ws.cell(row=n + 1, column=10, value=row[8])
            #Data pomiaru
            ws.cell(row=n + 1, column=11, value=row[9])
            #Numer foto
            ws.cell(row=n + 1, column=12, value=row[10])
            #Uwagi
            ws.cell(row=n + 1, column=13, value=row[13])

            for m in range(1, 14):
                try:
                    ws.cell(row=n + 1, column=m).style = stylest
                except:
                    ws.cell(row=n + 1, column=m).style = "standart"
            n+=1

        addScatterChart(ws, chartTitle, row[11], 7, 6, minRowOst, n+1)

    wb.save(os.path.join(folder, outputXlsx))
    wb.close()
    os.remove(os.path.join(folder, 'template.xlsx'))
    return

if __name__ == '__main__':
    fold_shp = r"C:\robo\_warstwy_tymczasowe\Zalacznik_1\w1\W1_DOP_1\shp"
    folder = r"C:\robo\_warstwy_tymczasowe\Zalacznik_1\w1\W1_DOP_1"
    liniowe = os.path.join(fold_shp, "W1_DOP_1_linie_przekrojow.shp")
    korytowe = os.path.join(fold_shp, "W1_DOP_1_przekroje_korytowe_1.shp")
    dolinowe = os.path.join(fold_shp, "W1_DOP_1_przekroje_terasy_1.shp")

    xlsx = 'W1_DOP_1_przekroje_dolinowe.xlsx'

    insertXlsxData(pktPrzekroi(liniowe, korytowe, dolinowe, 18), folder, xlsx)

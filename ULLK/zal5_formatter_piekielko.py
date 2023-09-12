#!/usr/bin/env python
#-*- coding: utf-8 -*-

import collections
import openpyxl
import pyexcel
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
import os

def styleHeadline():
    hstyle = NamedStyle(name="headline")
    hstyle.font = Font(bold=True, size=9)
    bd = Side(style='thin', color="000000")
    hstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    hstyle.fill = PatternFill("solid", fgColor="00C0C0C0")
    hstyle.alignment = Alignment(horizontal="center", vertical="center")
    return hstyle

def styleHeadline2():
    hstyle2 = NamedStyle(name="headline2")
    hstyle2.font = Font(bold=True, size=10)
    bd = Side(style='thin', color="000000")
    hstyle2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    hstyle2.fill = PatternFill("solid", fgColor="00CCCCFF")
    hstyle2.alignment = Alignment(horizontal="center", vertical="center")
    return hstyle2

def styleObr():
    obrstyle = NamedStyle(name="obreb")
    obrstyle.font = Font(bold=True, size=9)
    # bd = Side(style='thin', color="000000")
    # obrstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    obrstyle.fill = PatternFill("solid", fgColor="00808080")
    obrstyle.alignment = Alignment(horizontal="center", vertical="center")
    return obrstyle

def styleStandart():
    sstyle = NamedStyle(name="standart")
    sstyle.font = Font(bold=False, size=9)
    bd = Side(style='thin', color="000000")
    sstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sstyle.alignment=Alignment(horizontal="center", vertical="center")
    return sstyle

def styleStandart2():
    sstyle2 = NamedStyle(name="standart2")
    sstyle2.font = Font(bold=True, size=9)
    bd = Side(style='thin', color="000000")
    sstyle2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sstyle2.alignment=Alignment(horizontal="center", vertical="center")
    return sstyle2

def xls2xlsx(file):
    try:
        fname, ext = os.path.splitext(file)
        if os.path.isfile(file) and ext == '.xls':
            xlsx = fname + '.xlsx'
            pyexcel.save_book_as(file_name=file, dest_file_name=xlsx)
            return xlsx
        else:
            return file
    except Exception as e:
        raise e

def createTemplateXlsx(folder, output_xlsx):
    wb = openpyxl.Workbook()
    headlinestyle=styleHeadline()
    headlinestyle2=styleHeadline2()

    ws0 = wb.create_sheet('PKT I', 0)
    ws0['A1']='I.  DZIAŁKI OBJĘTE INWESTYCJĄ'
    ws0['A2'] = 'L.p.'
    ws0['B2'] = 'Numer działki'
    ws0['C2'] = 'Obręb'
    ws0['D2'] = 'Numer księgi wieczystej'
    ws0['E2'] = 'Oznaczenie właściciela/użytkownika wieczystego/ władającego/trwałego zarządcy'

    ws0.merge_cells('A1:E1')
    try:
        ws0['A1'].style = headlinestyle2
    except:
        ws0['A1'].style="headline2"

    for m in range(1, 6):
        try:
            ws0.cell(row=2, column=m).style=headlinestyle
        except:
            ws0.cell(row=2, column=m).style="headline"

    ws1=wb.create_sheet('PKT II', 1)
    ws1['A1'] = 'II.  DZIAŁKI POŁOŻONE W LINIACH ROZGRANICZAJĄCYCH TEREN INWESTYCJI BĘDĄCE WŁASNOŚCIĄ SKARBU PAŃSTWA ' \
                'W UŻYTKOWANIU WIECZYSTYM PKP S.A. ORAZ DZIAŁKI BĘDĄCE WŁASNOŚCIĄ PKP S.A. (NIE PODLEGAJĄCE PODZIAŁOWI)'
    ws1['A2'] = 'L.p.'
    ws1['B2'] = 'Numer działki'
    ws1['C2'] = 'Obręb'
    ws1['D2'] = 'Numer księgi wieczystej'
    ws1['E2'] = 'Powierzchnia działki [ha]'
    ws1['F2'] = 'Powierzchnia zajmowanej działki [ha]'
    ws1['G2'] = 'Lokalizacja względnem linii kolejowej [km]'
    ws1['H2'] = 'Oznaczenie właściciela/użytkownika wieczystego/ władającego/trwałego zarządcy'

    ws1.merge_cells('A1:H1')
    try:
        ws1['A1'].style=headlinestyle2
    except:
        ws1['A1'].style="headline2"

    for m in range(1, 9):
        try:
            ws1.cell(row=2, column=m).style=headlinestyle
        except:
            ws1.cell(row=2, column=m).style="headline"

    ws2 = wb.create_sheet('PKT III', 2)
    ws2['A1'] = 'III. DZIAŁKI POŁOŻONE W LINIACH ROZGRANICZAJĄCYCH TEREN INWESTYCJI, NIE BĘDĄCE WŁASNOŚCIĄ ALBO PRZEDMIOTEM UŻYTKOWANIA WIECZYSTEGO PKP S.A. LUB PKP PLK S.A. –' \
              '-PRZEZNACZONE POD INWESTYCJĘ Z UWZGLĘDNIENIEM DZIAŁEK POWSTAŁYCH W WYNIKU PODZIAŁU NIERUCHOMOŚCI (W NAWIASIE DZIAŁKA PRZED PODZIAŁEM), ' \
              'KTÓRE W CAŁOŚCI LUB W WYNIKU PODZIAŁU NIERUCHOMOŚCI STANOWIĆ BĘDĄ WŁASNOŚĆ SKARBU PAŃSTWA W UŻYTKOWANIU WIECZYSTYM PKP PLK S.A. (DO TRWAŁEGO ZAJĘCIA):'
    ws2['A2'] = 'L.p.'
    ws2['B2'] = 'Numer działki'
    ws2['C2'] = 'Obręb'
    ws2['D2'] = 'Numer księgi wieczystej'
    ws2['E2'] = 'Dotychczasowa powierzchnia działki [ha]'
    ws2['F2'] = 'Powierzchnia działki przejmowanej [ha]'
    ws2['G2'] = 'Powierzchnia działki pozostającej u dotychczasowego właściciela [ha]'
    ws2['H2'] = 'Lokalizacja względnem linii kolejowej [km]'
    ws2['I2'] = 'Oznaczenie właściciela/użytkownika wieczystego/ władającego/trwałego zarządcy'

    ws2.merge_cells('A1:I1')
    try:
        ws2['A1'].style=headlinestyle2
    except:
        ws2['A1'].style="headline2"

    for m in range(1, 10):
        try:
            ws2.cell(row=2, column=m).style=headlinestyle
        except:
            ws2.cell(row=2, column=m).style="headline"

    ws3=wb.create_sheet('PKT IV', 3)
    ws3['A1']='IV.  DZIAŁKI KTÓRE PRZEJDĄ W TRWAŁY ZARZĄD GENERALNEJ DYREKCJI DRÓG KRAJOWYCH I AUTOSTRAD W ZWIĄZKU Z BUDOWĄ DROGI ' \
              'KRAJOWEJ ZGODNIE Z ART. 9s UST. 3e USTAWY Z DNIA 28 MARCA 2003R. O TRANSPORCIE KOLEJOWYM (w nawiasie działka przed podziałem) '
    ws3['A2']='L.p.'
    ws3['B2']='Numer działki'
    ws3['C2']='Obręb'
    ws3['D2']='Numer księgi wieczystej'
    ws3['E2']='Dotychczasowa powierzchnia działki [ha]'
    ws3['F2']='Powierzchnia działki przekazywanej [ha]'
    ws3['G2']='Lokalizacja względnem linii kolejowej [km]'
    ws3['H2']='Oznaczenie właściciela/użytkownika wieczystego/ władającego/trwałego zarządcy'

    ws3.merge_cells('A1:H1')
    try:
        ws3['A1'].style=headlinestyle2
    except:
        ws3['A1'].style="headline2"

    for m in range(1, 9):
        try:
            ws3.cell(row=2, column=m).style=headlinestyle
        except:
            ws3.cell(row=2, column=m).style="headline"

    ws4=wb.create_sheet('PKT V', 4)
    ws4['A1']='V.  DZIAŁKI (części działek) POŁOŻONE W GRANICY TERENU NIEZBĘDNEGO, NA KTÓRYCH OKREŚLONO OGRANICZENIA W KORZYSTANIU Z NIERUCHOMOŚCI ' \
              'W CELU ZAPEWNIENIA PRAWA DO WEJŚCIA NA TEREN NIERUCHOMOŚCI DLA PROWADZENIA INWESTYCJI KOLEJOWEJ ' \
              'zgodnie z art. 9q, ust. 1, pkt 6 ustawy z dnia 28 marca 2003r. o transporcie kolejowym:'
    ws4['A2']='L.p.'
    ws4['B2']='Numer działki'
    ws4['C2']='Obręb'
    ws4['D2']='Powierzchnia działki [ha]'
    ws4['E2']='Powierzchnia zajmowanej działki z tytułu ograniczenia w korzystaniu z nieruchomości [ha]'
    ws4['F2']='Lokalizacja względnem linii kolejowej [km]'
    ws4['G2']='Zakres prac'

    ws4.merge_cells('A1:G1')
    try:
        ws4['A1'].style=headlinestyle2
    except:
        ws4['A1'].style="headline2"

    for m in range(1, 8):
        try:
            ws4.cell(row=2, column=m).style=headlinestyle
        except:
            ws4.cell(row=2, column=m).style="headline"

    ws5=wb.create_sheet('PKT VI', 5)
    ws5['A1']='VI.  DZIAŁKI (części działek) POŁOŻONE W GRANICY TERENU NIEZBĘDNEGO, DO NIEODPŁATNEGO ZAJĘCIA NA CZAS REALIZACJI INWESTYCJI STANOWIĄCE TEREN DRÓG PUBLICZNYCH'
    ws5['A2']='L.p.'
    ws5['B2']='Numer działki'
    ws5['C2']='Obręb'
    ws5['D2']='Powierzchnia działki [ha]'
    ws5['E2']='Powierzchnia działki drogowej zajmowanej na czas realizacji inwestycji [ha]'
    ws5['F2']='Lokalizacja względnem linii kolejowej [km]'
    ws5['G2']='Zakres prac'

    ws5.merge_cells('A1:G1')
    try:
        ws5['A1'].style=headlinestyle2
    except:
        ws5['A1'].style="headline2"

    for m in range(1, 8):
        try:
            ws5.cell(row=2, column=m).style=headlinestyle
        except:
            ws5.cell(row=2, column=m).style="headline"

    ws6=wb.create_sheet('PKT VII', 6)
    ws6['A1']='VII.  DZIAŁKI POŁOŻONE W GRANICY TERENU NIEZBĘDNEGO, DO NIEODPŁATNEGO ZAJĘCIA NA CZAS REALIZACJI INWESTYCJI STANOWIĄCE TEREN WÓD PŁYNĄCYCH'
    ws6['A2']='L.p.'
    ws6['B2']='Numer działki'
    ws6['C2']='Obręb'
    ws6['D2']='Powierzchnia działki [ha]'
    ws6['E2']='Powierzchnia terenu wód płynących zajmowane na czas realizacji inwestycji [ha]'
    ws6['F2']='Lokalizacja względnem linii kolejowej [km]'
    ws6['G2']='Zakres prac'

    ws6.merge_cells('A1:G1')
    try:
        ws6['A1'].style=headlinestyle2
    except:
        ws6['A1'].style="headline2"

    for m in range(1, 8):
        try:
            ws6.cell(row=2, column=m).style=headlinestyle
        except:
            ws6.cell(row=2, column=m).style="headline"

    ws7=wb.create_sheet('PKT VIII', 7)
    ws7['A1']='VIII.  DZIAŁKI NIE BĘDĄCE PRZEDMIOTEM WŁASNOŚCI ALBO UŻYTKOWANIA WIECZYSTEGO PKP S.A. LUB PKP PLK S.A., ' \
              'KTÓRE PO PODZIALE POZOSTANĄ W DOTYCHCZASOWYM UŻYTKOWANIU (w nawiasie działka przed podziałem)'
    ws7['A2']='L.p.'
    ws7['B2']='Numer działki'
    ws7['C2']='Obręb'
    ws7['D2']='Dotychczasowa powierzchnia działki [ha]'
    ws7['E2']='Powierzchnia działki pozostającej u dotychczasowego właściciela [ha]'
    ws7['F2']='Lokalizacja względnem linii kolejowej [km]'
    ws7['G2']='Oznaczenie właściciela/użytkownika wieczystego/ władającego/trwałego zarządcy'

    ws7.merge_cells('A1:G1')
    try:
        ws7['A1'].style=headlinestyle2
    except:
        ws7['A1'].style="headline2"

    for m in range(1, 8):
        try:
            ws7.cell(row=2, column=m).style=headlinestyle
        except:
            ws7.cell(row=2, column=m).style="headline"

    wb.save(os.path.join(folder, output_xlsx))
    wb.close()
    return output_xlsx

def insertObr(worksheet,dict, key, n, col_count):
    styleobreb=styleObr()
    section='Województwo: ' + dict[key][0] + ', Powiat: ' + dict[key][1] + ', Gmina: ' + dict[key][2] + ', Obręb: ' + \
            dict[key][3]
    worksheet.cell(row=n + 3, column=1, value=section)
    worksheet.merge_cells(start_row=n + 3, start_column=1, end_row=n + 3, end_column=col_count)
    merged_cell=worksheet.cell(row=n + 3, column=1)
    try:
        merged_cell.style=styleobreb
    except:
        merged_cell.style="obreb"
    return

def insertData(worksheet, dict, key, col_count, n, i):
    stylest=styleStandart()
    stylest2=styleStandart2()

    for k in range(1, col_count):
        if k == 1:
            worksheet.cell(row=n + 3, column=k, value=str(i) + '.')
        else:
            worksheet.cell(row=n + 3, column=k, value=dict[key][k + 2])

    for m in range(1, col_count):
        if m == 2:
            try:
                worksheet.cell(row=n + 3, column=m).style=stylest2
            except:
                worksheet.cell(row=n + 3, column=m).style="standart2"
        else:
            try:
                worksheet.cell(row=n + 3, column=m).style=stylest
            except:
                worksheet.cell(row=n + 3, column=m).style="standart"

    return

def rewriteDraftXls(folder, draft, template):

    try:
        wb = openpyxl.load_workbook(os.path.join(folder, draft))
    except:
        try:
            draft_xlsx = xls2xlsx(os.path.join(folder, draft))
            wb=openpyxl.load_workbook(os.path.join(folder, draft_xlsx))

        except Exception as e:
            raise e

    wss = wb['zajęcie stałe']
    wsc = wb['zajęcie czasowe']
    wb2 = openpyxl.load_workbook(os.path.join(folder, template))
    ws1 = wb2['PKT I']

    '''Uzupełnienie zakładki PKT I w arkusu wynikowym'''
    pktI = {}

    for row in wss.values:
        if row[1] != 'Identyfikator działki':
            if row[2] not in pktI:
                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-NR KW, 7- WLASCICIEL
                pktI[row[1]] = [row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], row[2], row[1][:13], row[12], row[11]]

    for row in wsc.values:
        if row[1] != 'Identyfikator działki':
            if row[2] not in pktI:
                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-NR KW, 7- WLASCICIEL
                pktI[row[1]] = [row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], row[2], row[1][:13], row[12], row[11]]

    # print(pktI)
    pktI = collections.OrderedDict(sorted(pktI.items()))

    obreb = ''
    i=1
    n=0
    for key in pktI.keys():
        if obreb != pktI[key][5]:
            obreb = pktI[key][5]
            insertObr(ws1, pktI, key, n, 5)

            n += 1
            insertData(ws1, pktI, key, 6, n, i)
        else:
            insertData(ws1, pktI, key, 6, n, i)

        i+=1
        n+=1

    '''Uzupełnienie zakładki PKT II w arkusu wynikowym'''
    pktII = {}
    ws2=wb2['PKT II']
    for row in wss.values:
        if row[1] != 'Identyfikator działki':
            if row[2] not in pktII:
                s=str(round(float(row[20]) / 10000, 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[25] == 'LK104 PROJ':
                    lok = row[24]+'\n'+'(LK 104 projektowana)'
                else:
                    lok = row[24]
                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-NR KW, 7-POW DZIALKI ha, 8-POW DZIALKI W INWESTYCJI ha,
                # 9-LOKALIZACJA, 10-WLASCICIEL
                if 'Nieruchomości będące w użytkowaniu wieczystym PKP S.A.' in row[23]:
                    pktII[row[1]] = [row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], row[2], row[1][:13], row[12], row[7], s.replace('.', ','), lok, row[11]]

    pktII=collections.OrderedDict(sorted(pktII.items()))
    obreb=''
    i=1
    n=0
    for key in pktII.keys():
        if obreb != pktII[key][5]:
            obreb=pktII[key][5]
            insertObr(ws2, pktII, key, n, 8)
            n+=1
            insertData(ws2, pktII, key, 9, n, i)
        else:
            insertData(ws2, pktII, key, 9, n, i)

        i+=1
        n+=1

    '''Uzupełnienie zakładki PKT III w arkusu wynikowym'''
    # sprawdzić bo wychodzi mi inna liczba działek niż PZT

    pktIII={}
    ws3=wb2['PKT III']
    for row in wss.values:
        if row[1] != 'Identyfikator działki':
            if 'PKP PLK S.A. nabywa prawo użytkowania wieczystego w myśl przepisu art. 9s ust. 3' in row[23]:
                s=str(round(float(row[20]) / 10000, 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[25] == 'LK104 PROJ':
                    lok=row[24] + '\n' + '(LK 104 projektowana)'
                else:
                    lok=row[24]
                if row[18] == 'TAK':
                    numery = row[19]+'\n'+'('+row[2]+')'
                else:
                    numery = row[2]
                pow_poz = str(round(float(row[7].replace(',','.')) - float(s),4))
                while len(pow_poz.split('.')[-1]) < 4:
                    pow_poz=pow_poz + '0'

                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-NR KW, 7-POW DZIALKI ha, 8-POW DZIALKI W INWESTYCJI ha,
                # 9-POW POZOSTALA, 10-LOKALIZACJA, 11-WLASCICIEL

                if row[1] not in pktIII:
                    pktIII[row[1]] = [row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                   row[1][:13], row[12], row[7], s.replace('.',','), pow_poz.replace('.',','), lok, row[11]]
                else:
                    pktIII[row[1]+'_id_'+str(row[0])] = [row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                    row[1][:13], row[12], row[7], s.replace('.', ','), pow_poz.replace('.', ','), lok,
                                    row[11]]

    pktIII=collections.OrderedDict(sorted(pktIII.items()))
    obreb=''
    i=1
    n=0
    for key in pktIII.keys():
        if obreb != pktIII[key][5]:
            obreb=pktIII[key][5]
            insertObr(ws3, pktIII, key, n, 9)
            n+=1
            insertData(ws3, pktIII, key, 10, n, i)
        else:
            insertData(ws3, pktIII, key, 10, n, i)

        i+=1
        n+=1

    '''Uzupełnienie zakładki PKT IV w arkusu wynikowym'''
    pktIV={}
    ws4=wb2['PKT IV']
    for row in wss.values:
        if row[1] != 'Identyfikator działki':
            if 'przejdą w trwały zarząd Generalnej Dyrekcji Dróg Krajowych i Autostrad w związku z budową drogi krajowej zgodnie z art. 9s ust. 3' in row[23]:
                s=str(round(float(row[20]) / 10000, 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[25] == 'LK104 PROJ':
                    lok=row[24] + '\n' + '(LK 104 projektowana)'
                else:
                    lok=row[24]
                if row[18] == 'TAK':
                    numery=row[19] + '\n' + '(' + row[2] + ')'
                else:
                    numery=row[2]

                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-NR KW, 7-POW DZIALKI ha, 8-POW DZIALKI W INWESTYCJI ha,
                # 9-LOKALIZACJA, 10-WLASCICIEL
                if row[1] not in pktIV:
                    pktIV[row[1]]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                row[1][:13], row[12], row[7], s.replace('.', ','), lok, row[11]]
                else:
                    pktIV[row[1]+'_id_'+str(row[0])]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                   row[1][:13], row[12], row[7], s.replace('.', ','), lok, row[11]]

    pktIV=collections.OrderedDict(sorted(pktIV.items()))
    obreb=''
    i=1
    n=0
    for key in pktIV.keys():
        if obreb != pktIV[key][5]:
            obreb=pktIV[key][5]
            insertObr(ws4, pktIV, key, n, 8)
            n+=1
            insertData(ws4, pktIV, key, 9, n, i)
        else:
            insertData(ws4, pktIV, key, 9, n, i)

        i+=1
        n+=1

    '''Uzupełnienie zakładki PKT V w arkusu wynikowym'''
    pktV={}
    ws5=wb2['PKT V']
    for row in wsc.values:
        if row[1] != 'Identyfikator działki':
            if 'ograniczenia w korzystaniu z nieruchomości w celu zapewnienia prawa do wejścia na teren nieruchomości w myśl przepisu art. 9q ust. 1 pkt 6   ustawy o transporcie kolejowym' in row[21]:
                s=str(round(float(row[19]) / 10000, 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[24] == 'LK104 PROJ':
                    lok=row[23] + '\n' + '(LK 104 projektowana)'
                else:
                    lok=row[23]

                if row[18] == 'TAK':
                    numery=row[20] + '\n' + '(' + row[2] + ')'
                else:
                    numery=row[2]

                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-POW DZIALKI ha, 7-POW DZIALKI W CZASOWYM ha,
                # 8-LOKALIZACJA, 9-ZAKRES PRAC
                if row[1] not in pktV:
                    pktV[row[1]]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                               row[1][:13], row[7], s.replace('.', ','), lok, row[22]]
                else:
                    pktV[row[1]+'_id_'+str(row[0])]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                  row[1][:13], row[7], s.replace('.', ','), lok, row[22]]

    pktV=collections.OrderedDict(sorted(pktV.items()))
    obreb=''
    i=1
    n=0
    for key in pktV.keys():
        if obreb != pktV[key][5]:
            obreb=pktV[key][5]
            insertObr(ws5, pktV, key, n, 7)
            n+=1
            insertData(ws5, pktV, key, 8, n, i)
        else:
            insertData(ws5, pktV, key, 8, n, i)

        i+=1
        n+=1

    '''Uzupełnienie zakładki PKT VI w arkusu wynikowym'''
    pktVI={}
    ws6=wb2['PKT VI']
    for row in wsc.values:
        if row[1] != 'Identyfikator działki':
            if 'pasy dróg publicznych' in row[21]:
                s=str(round(float(row[19]) / 10000, 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[24] == 'LK104 PROJ':
                    lok=row[23] + '\n' + '(LK 104 projektowana)'
                else:
                    lok=row[23]

                if row[18] == 'TAK':
                    numery=row[20] + '\n' + '(' + row[2] + ')'
                else:
                    numery=row[2]

                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-POW DZIALKI ha, 7-POW DZIALKI W CZASOWYM ha,
                # 8-LOKALIZACJA, 9-ZAKRES PRAC
                if row[1] not in pktVI:
                    pktVI[row[1]]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                               row[1][:13], row[7], s.replace('.', ','), lok, row[22]]
                else:
                    pktVI[row[1]+'_id_'+str(row[0])]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                   row[1][:13], row[7], s.replace('.', ','), lok, row[22]]

    pktVI=collections.OrderedDict(sorted(pktVI.items()))
    obreb=''
    i=1
    n=0
    for key in pktVI.keys():
        if obreb != pktVI[key][5]:
            obreb=pktVI[key][5]
            insertObr(ws6, pktVI, key, n, 7)
            n+=1
            insertData(ws6, pktVI, key, 8, n, i)
        else:
            insertData(ws6, pktVI, key, 8, n, i)

        i+=1
        n+=1

    '''Uzupełnienie zakładki PKT VII w arkusu wynikowym'''
    pktVII={}
    ws7=wb2['PKT VII']
    for row in wsc.values:
        if row[1] != 'Identyfikator działki':
            if 'wody płynące' in row[21] and '9ya ust. 1 ustawy o transporcie kolejowym' in row[21]:
                s=str(round(float(row[19]) / 10000, 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[24] == 'LK104 PROJ':
                    lok=row[23] + '\n' + '(LK 104 projektowana)'
                else:
                    lok=row[23]

                if row[18] == 'TAK':
                    numery=row[20] + '\n' + '(' + row[2] + ')'
                else:
                    numery=row[2]

                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-POW DZIALKI ha, 7-POW DZIALKI W CZASOWYM ha,
                # 8-LOKALIZACJA, 9-ZAKRES PRAC
                if row[1] not in pktVII:
                    pktVII[row[1]]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                               row[1][:13], row[7], s.replace('.', ','), lok, row[22]]
                else:
                    pktVII[row[1]+'_id_'+str(row[0])]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                    row[1][:13], row[7], s.replace('.', ','), lok, row[22]]


    pktVII=collections.OrderedDict(sorted(pktVII.items()))
    obreb=''
    i=1
    n=0
    for key in pktVII.keys():
        if obreb != pktVII[key][5]:
            obreb=pktVII[key][5]
            insertObr(ws7, pktVII, key, n, 7)
            n+=1
            insertData(ws7, pktVII, key, 8, n, i)
        else:
            insertData(ws7, pktVII, key, 8, n, i)

        i+=1
        n+=1

    '''Uzupełnienie zakładki PKT VIII w arkusu wynikowym'''
    pktVIII={}
    ws8=wb2['PKT VIII']
    for row in wss.values:
        if row[1] != 'Identyfikator działki':
            if row[18] == 'TAK' and row[21] is not None:
                s=str(round(float(row[7].replace(',', '.')) - (float(row[20]) / 10000), 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[25] == 'LK104 PROJ':
                    lok=row[24] + '\n' + '(LK 104 projektowana)'
                else:
                    lok=row[24]

                if row[18] == 'TAK':
                    numery=row[21] + '\n' + '(' + row[2] + ')'
                else:
                    numery=row[2]

                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-POW DZIALKI ha, 7-POW DZIALKI W CZASOWYM ha,
                # 8-LOKALIZACJA, 9-WLASCICIEL
                if row[1] not in pktVIII:
                    pktVIII[row[1]]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                               row[1][:13], row[7], s.replace('.', ','), lok, row[11]]
                else:
                    pktVIII[row[1]+'_id_'+str(row[0])]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                     row[1][:13], row[7], s.replace('.', ','), lok, row[11]]

            if row[22] is not None:
                s=str(round(float(row[7].replace(',', '.')) - (float(row[20]) / 10000), 4))
                while len(s.split('.')[-1]) < 4:
                    s=s + '0'
                if row[25] == 'LK104 PROJ':
                    lok=row[24] + '\n' + '(LK 104 projektowana)'
                else:
                    lok=row[24]

                if row[18] == 'TAK':
                    numery=row[22] + '\n' + '(' + row[2] + ')'
                else:
                    numery=row[2]

                # 0-WOJ, 1-POW, 2-GMINA, 3-NAZWA OBREBU, 4-NR DZIALKI, 5-NR OBREBU, 6-POW DZIALKI ha, 7-POW DZIALKI W CZASOWYM ha,
                # 8-LOKALIZACJA, 9-WLASCICIEL
                if row[1] not in pktVIII:
                    pktVIII[row[1]]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                row[1][:13], row[7], s.replace('.', ','), lok, row[11]]
                else:
                    pktVIII[row[1]+'_id_'+str(row[0])]=[row[3], row[4].replace('powiat ', ''), row[5], row[1][9:13] + ' ' + row[6], numery,
                                     row[1][:13], row[7], s.replace('.', ','), lok, row[11]]

    pktVIII=collections.OrderedDict(sorted(pktVIII.items()))
    obreb=''
    i=1
    n=0
    for key in pktVIII.keys():
        if obreb != pktVIII[key][5]:
            obreb=pktVIII[key][5]
            insertObr(ws8, pktVIII, key, n, 7)
            n+=1
            insertData(ws8, pktVIII, key, 8, n, i)
        else:
            insertData(ws8, pktVIII, key, 8, n, i)

        i+=1
        n+=1

    output_xlsx = 'Wykaz załacznik 5 - przetworzony wykaz GIS.xlsx'
    wb2.save(os.path.join(folder, output_xlsx))
    wb.close()
    wb2.close()
    os.remove(os.path.join(work_dir, template))
    return output_xlsx

if __name__ == '__main__':
    work_dir = input("Wklej sciezke do folderu z plikami xlsx:")
    draftxls = input("Nazwa pliku z wykazem z GIS (*xlsx, *xls):")
    rewriteDraftXls(work_dir, draftxls, createTemplateXlsx(work_dir, "template.xlsx"))

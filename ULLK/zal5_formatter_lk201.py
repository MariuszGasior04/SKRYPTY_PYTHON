#!/usr/bin/env python
#-*- coding: utf-8 -*-

import openpyxl
import pyexcel
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
import os

def styleHeadline():
    hstyle = NamedStyle(name="headline")
    hstyle.font = Font(bold=True, size=11)
    bd = Side(style='thin', color="000000")
    hstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    hstyle.fill = PatternFill("solid", fgColor="00C0C0C0")
    hstyle.alignment = Alignment(horizontal="center", vertical="center")
    return hstyle

def styleObr():
    obrstyle = NamedStyle(name="obreb")
    obrstyle.font = Font(bold=True, size=9)
    # bd = Side(style='thin', color="000000")
    # obrstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    obrstyle.fill = PatternFill("solid", fgColor="00808080")
    obrstyle.alignment = Alignment(horizontal="center", vertical="center")
    return obrstyle

def styleStandart():
    sstyle = NamedStyle(name="standart")
    sstyle.font = Font(bold=False, size=11)
    bd = Side(style='thin', color="000000")
    sstyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sstyle.alignment=Alignment(horizontal="center", vertical="center")
    return sstyle

def xls2xlsx(file):
    try:
        fname, ext = os.path.splitext(file)
        if os.path.isfile(file) and ext == '.xls':
            xlsx = fname + '.xlsx'
            pyexcel.save_book_as(file_name=file, dest_file_name=xlsx)
            return xlsx
        else:
            return file
    except Exception as e:
        raise e

def createTemplateXlsx(folder, output_xlsx):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('zał.5', 0)
    ws['A1'] = 'L.p.'
    ws['B1'] = 'Identyfikator działki'
    ws['C1'] = 'Obręb'
    ws['D1'] = 'Numer księgi wieczystej'
    ws['E1'] = 'Dotychczasowy właściciel'
    ws['F1'] = 'Numer działki przed podziałem'+'\n'+'(pow. w ha)'
    ws['G1'] = 'Numer działki pod inwestycję'+'\n'+'(pow. w ha)'
    ws['H1'] = 'Numer działki w gestii dotychczasowego właścielela'+'\n'+'(pow. w ha)'
    ws['I1'] = 'Numer działki pod inwestycję w myśl przepisu art. 9q ust. 1 pkt 6 (pow. w ha) /rodzaj ograniczenia '
    ws['J1'] = 'Nieruchomości stanowiące pasy dróg publicznych zajęte w ramach nieodpłatnego zajęcia na czas realizacji inwestycji w myśl przepisu 9ya ust. 1  (pow. w ha) /rodzaj ograniczenia '
    ws['K1'] = 'Nieruchomości stanowiące wody płynące zajęte w ramach nieodpłatnego zajęcia na czas realizacji inwestycji w myśl przepisu 9ya ust. 1  (pow. w ha) /rodzaj ograniczenia '
    ws['L1'] = 'Uwagi: podmiot wywłaszczenia na SP art. 9s ust. 3, 3b i 3e'

    headlinestyle = styleHeadline()
    for m in range(1, 13):
        try:
            ws.cell(row=1, column=m).style=headlinestyle
        except:
            ws.cell(row=1, column=m).style="headline"

    wb.save(os.path.join(folder, output_xlsx))
    wb.close()
    return output_xlsx

def insertData(worksheet, row, n):
    stylest=styleStandart()
    for m in range(1, 13):
        try:
            worksheet.cell(row=n + 1, column=m).style = stylest
        except:
            worksheet.cell(row=n + 1, column=m).style = "standart"

    worksheet.cell(row=n + 1, column=1, value=n)
    worksheet.cell(row=n + 1, column=2, value=row[1])
    worksheet.cell(row=n + 1, column=3, value=row[6])
    worksheet.cell(row=n + 1, column=4, value=row[12])
    worksheet.cell(row=n + 1, column=5, value=row[11])
    s0 = str(row[7])
    while len(s0.split('.')[-1]) < 4:
        s0=s0 + '0'
    nr_przed_podz=row[2] + '\n' + '(' +s0+ ')'
    worksheet.cell(row=n + 1, column=6, value=nr_przed_podz)

    if row[16] == 'CZASOWE' and row[17] == 'TAK':
        s1=str(round(float(row[20]) / 10000, 4))
        while len(s1.split('.')[-1]) < 4:
            s1 = s1 + '0'
        nr_po_podz=row[19] + ' (' + s1 + ')'
        worksheet.cell(row=n + 1, column=7, value=nr_po_podz)
        if row[22] == 'Nieruchomości stanowiące pasy dróg publicznych, zajęte w ramach nieodpłatnego zajęcia na czas realizacji inwestycji w myśl przepisu 9ya ust.1 ustawy o transporcie kolejowym':
            roboty=nr_po_podz + '\n' + row[23]
            worksheet.cell(row=n + 1, column=10, value=roboty)
        elif row[22] == 'Nieruchomości stanowiące wody płynące, zajęte w ramach nieodpłatnego zajęcia na czas realizacji inwestycji w myśl przepisu 9ya ust. 1 ustawy o transporcie kolejowym.':
            roboty=nr_po_podz + '\n' + row[23]
            worksheet.cell(row=n + 1, column=11, value=roboty)
        else:
            roboty=nr_po_podz + '\n' + row[23]
            worksheet.cell(row=n + 1, column=9, value=roboty)

    elif row[16] == 'STAŁE' and row[17] == 'TAK':
        s2=str(round(float(row[21]) / 10000, 4))
        while len(s2.split('.')[-1]) < 4:
            s2=s2 + '0'
        nr_po_podz=row[18] + ' (' + s2 + ')'
        worksheet.cell(row=n + 1, column=7, value=nr_po_podz)
        wywlaszczenie=row[18] + ' - PKP PLK S.A. - uż. wiecz.'
        worksheet.cell(row=n + 1, column=12, value=wywlaszczenie)
        s3 = str(round(float(row[7]) - (float(row[20]) / 10000), 4))
        while len(s3.split('.')[-1]) < 4:
            s3 = s3 + '0'
        pow_pozos=row[19] + ' (' + s3 + ')'
        worksheet.cell(row=n + 1, column=8, value=pow_pozos)
    elif row[16] == 'CZASOWE' and row[17] is None:
        s4=str(round(float(row[20]) / 10000, 4))
        while len(s4.split('.')[-1]) < 4:
            s4=s4 + '0'
        nr_po_podz=row[2] + ' (' + s4 + ')'
        worksheet.cell(row=n + 1, column=7, value=nr_po_podz)

        if row[22] == 'Nieruchomości stanowiące pasy dróg publicznych, zajęte w ramach nieodpłatnego zajęcia na czas realizacji inwestycji w myśl przepisu 9ya ust.1 ustawy o transporcie kolejowym':
            roboty=nr_po_podz + '\n' + row[23]
            worksheet.cell(row=n + 1, column=10, value=roboty)
        elif row[22] == 'Nieruchomości stanowiące wody płynące, zajęte w ramach nieodpłatnego zajęcia na czas realizacji inwestycji w myśl przepisu 9ya ust. 1 ustawy o transporcie kolejowym.':
            roboty=nr_po_podz + '\n' + row[23]
            worksheet.cell(row=n + 1, column=11, value=roboty)
        else:
            roboty=nr_po_podz + '\n' + row[23]
            worksheet.cell(row=n + 1, column=9, value=roboty)
    elif row[16] == 'STAŁE' and row[17] is None:
        s5=str(round(float(row[20]) / 10000, 4))
        while len(s5.split('.')[-1]) < 4:
            s5=s5 + '0'
        nr_po_podz=row[2] + ' (' + s5 + ')'
        worksheet.cell(row=n + 1, column=7, value=nr_po_podz)
        if row[22] == 'Nieruchomości, które stają się własnością Skarbu Państwa, co do których PKP PLK S.A. nabywa prawo użytkowania wieczystego w myśl przepisu art. 9s ust.3 pkt 1 i ust 3b ustawy o transporcie kolejowym':
            wywlaszczenie=row[2] + ' - PKP PLK S.A. - uż. wiecz.'
            worksheet.cell(row=n + 1, column=12, value=wywlaszczenie)
    return

def rewriteDraftXls(folder, draft, template):
    try:
        wb = openpyxl.load_workbook(os.path.join(folder, draft))
    except:
        try:
            draft_xlsx = xls2xlsx(os.path.join(folder, draft))
            wb=openpyxl.load_workbook(os.path.join(folder, draft_xlsx))

        except Exception as e:
            raise e

    ws = wb['Arkusz1']
    wb2 = openpyxl.load_workbook(os.path.join(folder, template))
    ws2 = wb2['zał.5']
    i = 0
    n = 0
    obreb = ''
    styleobreb = styleObr()

    for row in ws.values:
        if i > 0:
            if obreb != row[6]:
                obreb = row[6]
                section = 'Województwo: '+row[3]+', powiat: '+row[4]+', gmina: '+row[5]+', obreb: '+row[1][:13]+' '+row[6]
                # print(section)
                ws2.cell(row = n+1, column = 1, value = section)
                ws2.merge_cells(start_row = n+1, start_column = 1, end_row = n+1, end_column = 12)
                merged_cell = ws2.cell(row=n + 1, column=1)
                merged_cell.style = styleobreb
                n += 1
                insertData(ws2, row, n)

            else:
                insertData(ws2, row, n)
        i+=1
        n+=1

        # print(row)
    output_xlsx='Wykaz załacznik 5 - przetworzony wykaz GIS.xlsx'
    wb2.save(os.path.join(folder, output_xlsx))
    wb.close()
    wb2.close()
    os.remove(os.path.join(work_dir, template))
    return

def mergeCells(folder, filename):
    wb=openpyxl.load_workbook(os.path.join(folder, filename))
    ws=wb['zał.5']
    c = 0
    r = 0
    dict = {}
    for col in ws.columns:
        c+=1
        for cell in col:
            r+=1
            if r == 1:
                naz_kol = cell.value
            if naz_kol == 'Identyfikator działki' and cell.value is not None:
                if cell.value in dict:
                    dict[cell.value][2] = r
                    dict[cell.value][3] = c
                    dict[cell.value][4] = ws.cell(row=r, column=8).value

                else:
                    cell8 = ws.cell(row=r, column=8)
                    dict[cell.value] = [r, c, r, c, cell8.value]
        r = 0
    # print(dict)
    for value in dict.values():
        if value[0] != value[2]:
            for n in [1, 2, 3, 4, 5, 6, 9, 10, 11]:
                ws.merge_cells(start_row=value[0], start_column=n, end_row=value[2], end_column=n)

            cell8 = ws.cell(row=value[0], column=8)
            cell8.value = value[4]
            ws.merge_cells(start_row=value[0], start_column=8, end_row=value[2], end_column=8)


    wb.save(os.path.join(folder, 'Zalacznik_5_analiza_narzedzia_ost.xlsx'))
    wb.close()

    return

if __name__ == '__main__':
    work_dir = input("Wklej sciezke do folderu z plikami xlsx:")
    draftxls = input("Nazwa pliku z wykazem z GIS (*xlsx, *xls):")

    mergeCells(work_dir, rewriteDraftXls(work_dir, draftxls, createTemplateXlsx(work_dir, 'template.xlsx')))